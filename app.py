from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
import math
from io import BytesIO
import tempfile
from collections import defaultdict

app = Flask(__name__)
UPLOAD_FOLDER = tempfile.mkdtemp()
ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_dimensions(dim_str):
    """Estrae larghezza e lunghezza arrotondando in eccesso"""
    if not dim_str:
        return None, None
    match = re.search(r'(\d+(?:,\d+)?)\s*x\s*(\d+(?:,\d+)?)', str(dim_str))
    if match:
        width = float(match.group(1).replace(',', '.'))
        length = float(match.group(2).replace(',', '.'))
        width = math.ceil(width)
        length = math.ceil(length)
        return str(width), str(length)
    return None, None

def format_number_italian(num):
    """Formatta numeri con virgola italiana"""
    if num is None:
        return None
    try:
        val = float(str(num).replace(',', '.'))
        if val == int(val):
            return str(int(val))
        return str(val).replace('.', ',')
    except:
        return num

def parse_time(time_str):
    """Estrae minuti e secondi"""
    if not time_str:
        return None, None
    try:
        parts = str(time_str).split(':')
        if len(parts) >= 3:
            hours = int(parts[0])
            minutes = int(parts[1]) + hours * 60
            seconds = int(parts[2])
            return minutes, seconds
    except:
        pass
    return None, None

def organize_by_family(data):
    """Organizza dati per famiglia e assegna lettere padre-figlio"""
    families = defaultdict(list)

    for item in data:
        cod = item['codice']
        match = re.match(r'(.+?)(?:_\d+)?$', cod)
        base_cod = match.group(1) if match else cod
        families[base_cod].append(item)

    result = []
    current_letter = ord('A')

    for base_cod in sorted(families.keys()):
        family_items = families[base_cod]

        if len(family_items) == 1:
            # Singolo
            item = family_items[0]
            item['padre_figlio'] = None
            result.append(item)
        else:
            # Famiglia con figli
            letter = chr(current_letter)

            # Aggiungi figli PRIMA
            for item in sorted(family_items, key=lambda x: x['codice']):
                item['padre_figlio'] = letter
                result.append(item)

            # Aggiungi padre DOPO (se non esiste)
            parent_exists = any(item['codice'] == base_cod for item in family_items)

            if not parent_exists:
                parent_item = {
                    'file_pdf': family_items[0].get('file_pdf'),
                    'codice': base_cod,
                    'padre_figlio': letter,
                    'quantita': None,
                    'larghezza': None,
                    'lunghezza': None,
                    'spessore': None,
                    'peso': None,
                    'materiale': None,
                    'minuti': None,
                    'secondi': None
                }
                result.append(parent_item)

            current_letter += 1

        # Riga vuota tra famiglie
        if base_cod != sorted(families.keys())[-1]:
            result.append(None)

    return result

def extract_raw_data(pdf_file, filename):
    """Estrae i dati grezzi dal PDF - VERSIONE SEMPLICE E AFFIDABILE"""
    data = []

    with pdfplumber.open(pdf_file) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text()

    # Cerca tutti i blocchi che iniziano con "Cod."
    # Pattern: Cod. CODICE Q.tà QTY ... Materiale MAT Spessore SPESS ... Peso PESO ... Dimensioni DIM ... Tempo TIME
    pattern = r'Cod\.\s+([^\s]+)\s+Q\.tà\s+(\d+)\s+.*?Materiale\s+([^\s]+)\s+Spessore\s+([\d,]+)\s*mm.*?Peso\s+([\d,]+)\s*Kg.*?Dimensioni\s+([^\s]+).*?Tempo\s+([0-9:]+)'

    matches = re.finditer(pattern, full_text, re.DOTALL)

    for match in matches:
        cod = match.group(1)
        qty = match.group(2)
        material = match.group(3)
        spessore = match.group(4).replace(',', '.')
        peso = match.group(5)
        dimensions = match.group(6)
        time_str = match.group(7)

        width, length = parse_dimensions(dimensions)
        minutes, seconds = parse_time(time_str)

        data.append({
            'file_pdf': filename,
            'codice': cod,
            'quantita': qty,
            'larghezza': width,
            'lunghezza': length,
            'spessore': spessore,
            'peso': peso,
            'materiale': material,
            'minuti': minutes,
            'secondi': seconds
        })

    return data

def create_excel(data):
    """Crea Excel dai dati organizzati"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taglio"

    headers = ['File PDF', 'Codice', 'Padre Figlio', 'Quantità', 'Larghezza', 'Lunghezza', 'Spessore', 'Peso', 'Materiale', 'Minuti', 'Secondi']
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in data:
        if row is None:
            ws.append([])
        else:
            ws.append([
                row.get('file_pdf'),
                row.get('codice'),
                row.get('padre_figlio'),
                row.get('quantita'),
                format_number_italian(row.get('larghezza')),
                format_number_italian(row.get('lunghezza')),
                format_number_italian(row.get('spessore')),
                format_number_italian(row.get('peso')),
                row.get('materiale'),
                row.get('minuti'),
                row.get('secondi')
            ])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    return wb

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/api/upload', methods=['POST'])
def upload():
    """Riceve PDF, estrae dati, ritorna UNICO Excel"""
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'No files selected'}), 400

    all_raw_data = []

    try:
        # Processa ogni PDF
        for file in files:
            if file and allowed_file(file.filename):
                try:
                    raw_data = extract_raw_data(file, file.filename)
                    all_raw_data.extend(raw_data)
                    print(f"✓ {file.filename}: {len(raw_data)} articoli")
                except Exception as e:
                    print(f"✗ {file.filename}: {str(e)}")
                    return jsonify({'error': f'Errore in {file.filename}: {str(e)}'}), 400
            else:
                return jsonify({'error': f'{file.filename} non è un PDF valido'}), 400

        if not all_raw_data:
            return jsonify({'error': 'Nessun dato estratto dai PDF'}), 400

        # Organizza TUTTI i dati UNA SOLA VOLTA
        organized_data = organize_by_family(all_raw_data)

        # Crea Excel
        wb = create_excel(organized_data)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='taglio_estratto.xlsx'
        )

    except Exception as e:
        print(f"Errore generale: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
